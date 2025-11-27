#!/usr/bin/env python3
"""
Generate comprehensive test report from pytest results.
This script parses JUnit XML and coverage XML to create a detailed report.
"""

import json
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def parse_junit_xml(junit_path):
    """Parse JUnit XML test results."""
    tree = ET.parse(junit_path)
    root = tree.getroot()

    tests = int(root.get("tests", 0))
    failures = int(root.get("failures", 0))
    errors = int(root.get("errors", 0))
    skipped = int(root.get("skipped", 0))
    time = float(root.get("time", 0))

    passed = tests - failures - errors - skipped

    failed_tests = []
    all_tests = []
    for testcase in root.findall(".//testcase"):
        classname = testcase.get("classname", "Unknown")
        name = testcase.get("name", "Unknown")
        time_taken = float(testcase.get("time", 0))

        # Determine status
        if testcase.find("failure") is not None:
            status = "failed"
            message = testcase.find("failure").get("message", "No message")
            failed_tests.append({"class": classname, "name": name, "message": message})
        elif testcase.find("error") is not None:
            status = "error"
            message = testcase.find("error").get("message", "No message")
            failed_tests.append({"class": classname, "name": name, "message": message})
        elif testcase.find("skipped") is not None:
            status = "skipped"
            message = ""
        else:
            status = "passed"
            message = ""

        all_tests.append({"class": classname, "name": name, "status": status, "time": time_taken, "message": message})

    return {"total": tests, "passed": passed, "failed": failures, "errors": errors, "skipped": skipped, "time": time, "failed_tests": failed_tests, "all_tests": all_tests}


def parse_coverage_xml(coverage_path):
    """Parse coverage XML report."""
    tree = ET.parse(coverage_path)
    root = tree.getroot()

    line_rate = float(root.get("line-rate", 0)) * 100
    branch_rate = float(root.get("branch-rate", 0)) * 100

    packages = []
    for package in root.findall(".//package"):
        pkg_name = package.get("name", "Unknown")
        pkg_line_rate = float(package.get("line-rate", 0)) * 100
        pkg_branch_rate = float(package.get("branch-rate", 0)) * 100

        packages.append({"name": pkg_name, "line_rate": pkg_line_rate, "branch_rate": pkg_branch_rate})

    return {"line_rate": line_rate, "branch_rate": branch_rate, "packages": packages}


def generate_markdown_report(junit_data, coverage_data):
    """Generate markdown report."""
    report = []
    report.append("# Test Execution Report\n")
    report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}\n")

    # Test Results Summary
    report.append("## Test Results Summary\n")
    report.append("| Metric | Count |")
    report.append("|--------|-------|")
    report.append(f"| Total Tests | {junit_data['total']} |")
    report.append(f"| ✅ Passed | {junit_data['passed']} |")
    report.append(f"| ❌ Failed | {junit_data['failed']} |")
    report.append(f"| ⚠️ Errors | {junit_data['errors']} |")
    report.append(f"| ⏭️ Skipped | {junit_data['skipped']} |")
    report.append(f"| ⏱️ Duration | {junit_data['time']:.2f}s |")
    report.append("")

    # Coverage Summary
    report.append("## Coverage Summary\n")
    report.append("| Metric | Coverage |")
    report.append("|--------|----------|")
    report.append(f"| Line Coverage | {coverage_data['line_rate']:.1f}% |")
    report.append(f"| Branch Coverage | {coverage_data['branch_rate']:.1f}% |")
    report.append("")

    # Coverage by Package
    if coverage_data["packages"]:
        report.append("## Coverage by Package\n")
        report.append("| Package | Line Rate | Branch Rate |")
        report.append("|---------|-----------|-------------|")
        for pkg in sorted(coverage_data["packages"], key=lambda x: x["line_rate"], reverse=True):
            report.append(f"| {pkg['name']} | {pkg['line_rate']:.1f}% | {pkg['branch_rate']:.1f}% |")
        report.append("")

    # Failed Tests Details
    if junit_data["failed_tests"]:
        report.append("## Failed Tests\n")
        for i, test in enumerate(junit_data["failed_tests"], 1):
            report.append(f"### {i}. {test['class']}::{test['name']}\n")
            report.append(f"```\n{test['message']}\n```\n")

    # Status
    report.append("## Status\n")
    if junit_data["failed"] == 0 and junit_data["errors"] == 0:
        report.append("✅ **All tests passed!**\n")
    else:
        report.append(f"❌ **{junit_data['failed'] + junit_data['errors']} test(s) failed**\n")

    if coverage_data["line_rate"] >= 80:
        report.append(f"✅ **Coverage meets threshold (>80%): {coverage_data['line_rate']:.1f}%**\n")
    else:
        report.append(f"⚠️ **Coverage below threshold (<80%): {coverage_data['line_rate']:.1f}%**\n")

    return "\n".join(report)


def generate_json_report(junit_data, coverage_data):
    """Generate JSON report."""
    return {"timestamp": datetime.now().isoformat(), "tests": junit_data, "coverage": coverage_data}


def generate_excel_report(junit_data, coverage_data):
    """Generate Excel report with multiple sheets."""
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not available, skipping Excel report generation", file=sys.stderr)
        return None

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal="center")

    # Summary headers
    ws_summary["A1"] = "Test Execution Report"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws_summary["A2"].font = Font(italic=True)

    # Test Results Summary
    ws_summary["A4"] = "Test Results Summary"
    ws_summary["A4"].font = Font(bold=True, size=14)

    headers = ["Metric", "Count"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    metrics = [("Total Tests", junit_data["total"]), ("Passed", junit_data["passed"]), ("Failed", junit_data["failed"]), ("Errors", junit_data["errors"]), ("Skipped", junit_data["skipped"]), ("Duration (s)", f"{junit_data['time']:.2f}")]

    for row, (metric, value) in enumerate(metrics, 6):
        ws_summary.cell(row=row, column=1).value = metric
        ws_summary.cell(row=row, column=2).value = value

    # Coverage Summary
    ws_summary["A13"] = "Coverage Summary"
    ws_summary["A13"].font = Font(bold=True, size=14)

    headers = ["Metric", "Coverage"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=14, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    coverage_metrics = [("Line Coverage", f"{coverage_data['line_rate']:.1f}%"), ("Branch Coverage", f"{coverage_data['branch_rate']:.1f}%")]

    for row, (metric, value) in enumerate(coverage_metrics, 15):
        ws_summary.cell(row=row, column=1).value = metric
        ws_summary.cell(row=row, column=2).value = value

    # Status
    ws_summary["A18"] = "Status"
    ws_summary["A18"].font = Font(bold=True, size=14)

    status_row = 19
    if junit_data["failed"] == 0 and junit_data["errors"] == 0:
        ws_summary.cell(row=status_row, column=1).value = "✅ All tests passed!"
        status_row += 1
    else:
        ws_summary.cell(row=status_row, column=1).value = f"❌ {junit_data['failed'] + junit_data['errors']} test(s) failed"
        status_row += 1

    if coverage_data["line_rate"] >= 80:
        ws_summary.cell(row=status_row, column=1).value = f"✅ Coverage meets threshold (>80%): {coverage_data['line_rate']:.1f}%"
    else:
        ws_summary.cell(row=status_row, column=1).value = f"⚠️ Coverage below threshold (<80%): {coverage_data['line_rate']:.1f}%"

    # Auto-adjust column widths for summary
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws_summary.column_dimensions[column_letter].width = adjusted_width

    # Tests sheet
    ws_tests = wb.create_sheet("All Tests")

    test_headers = ["Class", "Name", "Status", "Time (s)", "Message"]
    for col, header in enumerate(test_headers, 1):
        cell = ws_tests.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row, test in enumerate(junit_data["all_tests"], 2):
        ws_tests.cell(row=row, column=1).value = test["class"]
        ws_tests.cell(row=row, column=2).value = test["name"]
        ws_tests.cell(row=row, column=3).value = test["status"]
        ws_tests.cell(row=row, column=4).value = test["time"]
        ws_tests.cell(row=row, column=5).value = test["message"]

        # Color code status
        status_cell = ws_tests.cell(row=row, column=3)
        if test["status"] == "passed":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif test["status"] == "failed":
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif test["status"] == "error":
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif test["status"] == "skipped":
            status_cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # Auto-adjust column widths for tests
    for column in ws_tests.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for message column
        ws_tests.column_dimensions[column_letter].width = adjusted_width

    # Coverage sheet
    ws_coverage = wb.create_sheet("Coverage")

    coverage_headers = ["Package", "Line Rate (%)", "Branch Rate (%)"]
    for col, header in enumerate(coverage_headers, 1):
        cell = ws_coverage.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row, pkg in enumerate(sorted(coverage_data["packages"], key=lambda x: x["line_rate"], reverse=True), 2):
        ws_coverage.cell(row=row, column=1).value = pkg["name"]
        ws_coverage.cell(row=row, column=2).value = pkg["line_rate"]
        ws_coverage.cell(row=row, column=3).value = pkg["branch_rate"]

    # Auto-adjust column widths for coverage
    for column in ws_coverage.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws_coverage.column_dimensions[column_letter].width = adjusted_width

    # Failed Tests sheet (if any)
    if junit_data["failed_tests"]:
        ws_failed = wb.create_sheet("Failed Tests")

        failed_headers = ["#", "Class", "Name", "Message"]
        for col, header in enumerate(failed_headers, 1):
            cell = ws_failed.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row, test in enumerate(junit_data["failed_tests"], 2):
            ws_failed.cell(row=row, column=1).value = row - 1
            ws_failed.cell(row=row, column=2).value = test["class"]
            ws_failed.cell(row=row, column=3).value = test["name"]
            ws_failed.cell(row=row, column=4).value = test["message"]

        # Auto-adjust column widths for failed tests
        for column in ws_failed.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)  # Allow wider for message
            ws_failed.column_dimensions[column_letter].width = adjusted_width

    return wb


def main():
    """Main entry point."""
    junit_path = Path("coverage_reports/junit.xml")
    coverage_path = Path("coverage_reports/coverage.xml")

    if not junit_path.exists():
        print(f"Error: {junit_path} not found", file=sys.stderr)
        return 1

    if not coverage_path.exists():
        print(f"Error: {coverage_path} not found", file=sys.stderr)
        return 1

    # Parse reports
    junit_data = parse_junit_xml(junit_path)
    coverage_data = parse_coverage_xml(coverage_path)

    # Generate reports
    markdown_report = generate_markdown_report(junit_data, coverage_data)
    json_report = generate_json_report(junit_data, coverage_data)
    excel_workbook = generate_excel_report(junit_data, coverage_data)

    # Write markdown report
    report_path = Path("coverage_reports/TEST_REPORT.md")
    report_path.write_text(markdown_report)
    print(f"✅ Markdown report written to {report_path}")

    # Write JSON report
    json_path = Path("coverage_reports/test_report.json")
    json_path.write_text(json.dumps(json_report, indent=2))
    print(f"✅ JSON report written to {json_path}")

    # Write Excel report
    if excel_workbook:
        excel_path = Path("coverage_reports/test_report.xlsx")
        excel_workbook.save(excel_path)
        print(f"✅ Excel report written to {excel_path}")
    else:
        print("⚠️ Excel report not generated (openpyxl not available)")

    # Print summary to stdout
    print("\n" + markdown_report)

    # Return exit code based on test results
    if junit_data["failed"] > 0 or junit_data["errors"] > 0:
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
